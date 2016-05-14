from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Color, PatternFill, Style
from os.path import join
from subprocess import Popen, PIPE
from json import load, dump
from datetime import datetime
from socket import gethostname


def tcl_runner(the_script, split_newline=True):
    """
    Execute given script and wait for output.
    :param the_script: script path.
    :param split_newline: decide if to remove new line and return one string, or split new line and return list.
    :return: string if split_new line False, list if True
    """
    proc = Popen(['tclsh', the_script], stdout=PIPE, shell=True)
    out = str(proc.communicate()[0])
    if split_newline:
        out = out.replace('b\'', '').replace('\\r', '').split('\\n')
    else:
        out = out.replace('b\'', '').replace('\\r', '').replace('\\n', '')
    return out


def get_unit_name():
    """
    Read unit name (ULT Tag)
    :return: name as string
    """
    return tcl_runner(data['script_name'], split_newline=False).replace('ULTTAG: ', '').replace('\'', '')


def new_document(workbook, ws_title):
    """
    Create new workbook in excel document, defaults are "number", "ULT", "pass", "runtime", and a cell for each test to run.
    Also creates a 'map' file for the location of cells for pass/fail.
    :param ws_title: worksheet title
    :param workbook: workbook object
    :return: new worksheet
    """
    global start_cell, unit_details, data
    cell_map = {'pass': {}, 'fail': {}}
    # set worksheet
    worksheet = workbook.create_sheet(title=ws_title, index=0)
    # Create Pass column and merge, merge is dependent on amount of tests
    worksheet.merge_cells('%s1:%s1' % (start_cell, chr(ord(start_cell) + len(unit_details) + len(data['tests']) - 1)))
    worksheet['%s1' % start_cell] = 'PASS'
    worksheet['%s1' % start_cell].style = Style(fill=PatternFill(patternType='solid', fgColor=Color('007F00')))
    worksheet['%s1' % start_cell].alignment = Alignment(horizontal='center')
    cell_map['pass'] = {}
    cell_map['fail'] = {}
    for detail in unit_details:
        cell_map['pass'].update({detail: {'location': '%s' % start_cell, 'longest_string': len(detail)}})
        worksheet['%s2' % start_cell] = detail
        worksheet['%s2' % start_cell].alignment = Alignment(horizontal='center')
        worksheet.column_dimensions[start_cell].width = len(detail) * 1.2
        start_cell = chr(ord(start_cell) + 1)
    for test in sorted(data['tests']):
        cell_map['pass'].update({test: {'location': '%s' % start_cell, 'longest_string': len(test)}})
        worksheet['%s2' % start_cell] = test
        worksheet['%s2' % start_cell].alignment = Alignment(horizontal='center')
        worksheet.column_dimensions[start_cell].width = len(test) * 1.2
        start_cell = chr(ord(start_cell) + 1)
    # Move start cell for Fail columns
    start_cell = chr(64 + len(unit_details) + len(data['tests']) + 2)
    # Create Fail column and merge, merge is dependent on amount of tests
    worksheet.merge_cells('%s1:%s1' % (start_cell, chr(ord(start_cell) + len(unit_details) + len(data['tests']) - 1)))
    worksheet['%s1' % start_cell] = 'Fail'
    worksheet['%s1' % start_cell].style = Style(fill=PatternFill(patternType='solid', fgColor=Color('FFFF0000')))
    worksheet['%s1' % start_cell].alignment = Alignment(horizontal='center')
    for detail in unit_details:
        cell_map['fail'].update({detail: {'location': '%s' % start_cell, 'longest_string': len(detail)}})
        worksheet['%s2' % start_cell] = detail
        worksheet['%s2' % start_cell].alignment = Alignment(horizontal='center')
        worksheet.column_dimensions[start_cell].width = len(detail) * 1.2
        start_cell = chr(ord(start_cell) + 1)
    for test in sorted(data['tests']):
        cell_map['fail'].update({test: {'location': '%s' % start_cell, 'longest_string': len(test)}})
        worksheet['%s2' % start_cell] = test
        worksheet['%s2' % start_cell].alignment = Alignment(horizontal='center')
        worksheet.column_dimensions[start_cell].width = len(test) * 1.2
        start_cell = chr(ord(start_cell) + 1)

    with open(join(data['excel_path'], '%s_cell_map.json' % data['unit']), 'w') as out_map:
        dump(cell_map, out_map, sort_keys=True, indent=4, ensure_ascii=False)

    return worksheet


def main():
    """
    Main program
    :return: none
    """
    # Basic unit details and test results
    unit_results = {'ULT Tag': get_unit_name(),
                    'number': '',
                    'hostname': gethostname(),
                    'pass': '',
                    'runtime': '',
                    'tests': {test: '' for test in data['tests']}
                    }
    start_time = datetime.now()
    # Run tests and capture results
    for test in data['tests']:
        print('Running test %s' % test)
        test_log = tcl_runner(test)
        for log in test_log:
            print(log)
        unit_results['tests'][test] = 'pass' if 'pass' in test_log else 'fail'
        print('End test %s, result: %s' % (test, unit_results['tests'][test]))
    if all(unit_results['tests'][test] == 'pass' for test in unit_results['tests']):
        unit_results['pass'] = 'pass'
    else:
        unit_results['pass'] = 'fail'
    unit_results['runtime'] = str(datetime.now() - start_time)
    # Open/Create main document
    try:
        excel_path = join(data['excel_path'], datetime.utcnow().strftime('%d-%m-%y') + '.xlsx')
        wb = load_workbook(excel_path)  # load existing document
        '''
        if all(data['unit'] is not sheet for sheet in wb):
            ws = new_document(wb, data['unit'])
        else:
            ws = wb[data['unit']]
        '''
        try:
            ws = wb[data['unit']]
        except KeyError:
            ws = new_document(wb, data['unit'])
        print('Opened existing document')
    except FileNotFoundError:  # create new document if none is present
        wb = Workbook()
        ws = new_document(wb, data['unit'])
        print('Created new document')
    # Load map file (cell locations)
    while True:  # Wait in case file not yet ready
        try:
            with open(join(data['excel_path'], '%s_cell_map.json' % data['unit']), 'r') as in_map:
                cell_map = load(in_map)
                print('Created map file')
            break
        except FileNotFoundError:
            pass
    n = 3
    print('Saving results')
    while True:
        # Find empty cell
        if ws['%s%d' % (cell_map[unit_results['pass']]['number']['location'], n)].value is None:
            unit_results['number'] = n-2
            for result in unit_results:
                # Skip tests
                if result is 'tests':
                    continue
                # Log results to cell
                ws['%s%d' % (cell_map[unit_results['pass']][result]['location'], n)] = unit_results[result]
                # Check if new string is longer then others, resize and save new width
                if len(str(unit_results[result])) > cell_map[unit_results['pass']][result]['longest_string']:
                    ws.column_dimensions[cell_map[unit_results['pass']][result]['location']].width = int(len(str(unit_results[result])) * 1.20)
                    cell_map[unit_results['pass']][result]['longest_string'] = len(str(unit_results[result]))
                ws['%s%d' % (cell_map[unit_results['pass']][result]['location'], n)].alignment = Alignment(horizontal='center')
            for test in data['tests']:
                # Log result to cell:
                ws['%s%d' % (cell_map[unit_results['pass']][test]['location'], n)] = unit_results['tests'][test]
                # Check if new string is longer then others, resize and save new width
                if len(str(unit_results['tests'][test])) > cell_map[unit_results['pass']][test]['longest_string']:
                    ws.column_dimensions[cell_map[unit_results['pass']][test]['location']].width = int(len(str(unit_results[test])) * 1.20)
                    cell_map[unit_results['pass']][test]['longest_string'] = len(str(unit_results[test]))
                ws['%s%d' % (cell_map[unit_results['pass']][test]['location'], n)].alignment = Alignment(horizontal='center')
            break
        else:
            n += 1
    # Save file
    wb.save(filename=join(data['excel_path'], datetime.utcnow().strftime('%d-%m-%y') + '.xlsx'))
    # Print results to user
    print('\n\n\nAll tests are done!')
    for test in data['tests']:
        print('%s result: %s' % (test, unit_results['tests'][test]))
    print('\n Unit is %s' % unit_results['pass'])
    print('Press enter to exit')
    input()
    return 1


if __name__ == '__main__':
    # Load preset data
    with open(r'.\data.json') as infile:  # load save file
        data = load(infile)
    # Set start cell and default details
    start_cell = 'A'
    unit_details = ('number', 'hostname', 'ULT Tag', 'pass', 'runtime',)
    # Run main
    main()
